#!/usr/bin/env python3
"""Import February 2026 diary attendance rows into the HRPayroll database.

This script reads the workbook, keeps only the primary sheet per employee,
and upserts the explicit work/OT values from the diary report into SLE_Attendances.

Usage:
  HRPAYROLL_CONNECTION_STRING="..." python3 scripts/import_feb2026_attendance.py \
    --workbook "/path/to/FEB2026_ALL_WORKERS_DiaryReports_1.xlsx"

The script supports --dry-run and will print a summary before modifying data.
"""

from __future__ import annotations

import argparse
import calendar
import os
import re
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Iterable, Optional

import pyodbc
from openpyxl import load_workbook


PRIMARY_SHEETS = {
    "Myint Than",
    "Naing Myint",
    "Chua Zi Jian",
    "Kyaw Zin Hein",
    "Sai Myat Soe",
    "Goh Woo Hang",
    "Dinh Van Cu",
    "Zhang Ji Jun",
}

TIME_RE = re.compile(r"(\d{1,2}:\d{2}\s*[AP]M)", re.IGNORECASE)
NUMBER_RE = re.compile(r"(?<!\w)(\d+(?:\.\d+)?)")


@dataclass(frozen=True)
class AttendanceRow:
    employee_id: int
    date: datetime.date
    status: str
    start: Optional[time]
    end: Optional[time]
    work_hours: float
    ot_hours: float
    site_project: Optional[str]
    transport: Optional[str]
    remarks: Optional[str]


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def parse_time(value: object) -> Optional[time]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value

    text = str(value).strip().upper().replace("．", ".").replace(".", "")
    text = re.sub(r"\s+", " ", text)
    if not text or text in {"-", "—"}:
        return None

    match = TIME_RE.search(text)
    if not match:
        return None

    token = match.group(1).replace(" ", "").upper()
    return datetime.strptime(token, "%I:%M%p").time()


def parse_number(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("$", "").replace(",", "")
    if not text or text in {"-", "—"}:
        return default

    match = NUMBER_RE.search(text)
    return float(match.group(1)) if match else default


def extract_times_and_numbers(value: object) -> tuple[Optional[time], Optional[time], Optional[float], Optional[float]]:
    if not isinstance(value, str):
        return None, None, None, None

    text = value.strip()
    times = TIME_RE.findall(text)
    nums = NUMBER_RE.findall(text)

    start = parse_time(times[0]) if len(times) >= 1 else None
    end = parse_time(times[1]) if len(times) >= 2 else None
    work = float(nums[0]) if len(nums) >= 1 else None
    ot = float(nums[1]) if len(nums) >= 2 else None
    return start, end, work, ot


def build_records(workbook_path: Path, employee_map: dict[str, int]) -> list[AttendanceRow]:
    wb = load_workbook(workbook_path, data_only=True)
    records: list[AttendanceRow] = []
    month_days = calendar.monthrange(2026, 2)[1]

    for sheet_name in wb.sheetnames:
        if sheet_name not in PRIMARY_SHEETS:
            continue

        ws = wb[sheet_name]
        employee_id = employee_map.get(normalize_text(sheet_name))
        if employee_id is None:
            raise RuntimeError(f"No employee found in database for sheet '{sheet_name}'")

        day_map: dict[int, AttendanceRow] = {}
        for row_idx in range(10, 38):
            day_value = ws.cell(row_idx, 3).value
            if day_value in (None, ""):
                continue

            day = int(day_value)
            date_value = datetime(2026, 2, day).date()
            raw_start = ws.cell(row_idx, 4).value
            raw_end = ws.cell(row_idx, 5).value
            raw_work = ws.cell(row_idx, 6).value
            raw_ot = ws.cell(row_idx, 7).value
            raw_site = ws.cell(row_idx, 8).value
            raw_transport = ws.cell(row_idx, 12).value

            text = " ".join(filter(None, [normalize_text(raw_start), normalize_text(raw_end), normalize_text(raw_site)]))
            start_time = parse_time(raw_start)
            end_time = parse_time(raw_end)
            work_hours = parse_number(raw_work, 0.0)
            ot_hours = parse_number(raw_ot, 0.0)
            remarks: Optional[str] = None

            if "MC" in text:
                status = "Leave"
                remarks = "MC"
                start_time = None
                end_time = None
                work_hours = 0.0
                ot_hours = 0.0
            elif "PUBLIC HOLIDAY" in text or re.search(r"\bPH\b", text):
                status = "Holiday"
                remarks = "CNY PUBLIC HOLIDAY" if "CNY" in text else "PUBLIC HOLIDAY"
                start_time = None
                end_time = None
                work_hours = 0.0
                ot_hours = 0.0
            else:
                status = "Present"
                if (start_time is None or end_time is None) and isinstance(raw_start, str):
                    parsed_start, parsed_end, parsed_work, parsed_ot = extract_times_and_numbers(raw_start)
                    start_time = start_time or parsed_start
                    end_time = end_time or parsed_end
                    if work_hours == 0.0 and parsed_work is not None:
                        work_hours = parsed_work
                    if ot_hours == 0.0 and parsed_ot is not None:
                        ot_hours = parsed_ot

            transport = "MRT" if raw_transport not in (None, "", " ", "-", "—") else None
            site_project = str(raw_site).strip() if raw_site not in (None, "", " ") else None

            if status == "Present" and start_time is None and end_time is None and work_hours == 0.0 and ot_hours == 0.0:
                # Skip stray blank rows that do not contain usable attendance data.
                continue

            day_map[day] = AttendanceRow(
                employee_id=employee_id,
                date=date_value,
                status=status,
                start=start_time,
                end=end_time,
                work_hours=round(work_hours, 2),
                ot_hours=round(ot_hours, 2),
                site_project=site_project,
                transport=transport,
                remarks=remarks,
            )

        # Any missing diary day in the workbook is treated as an absent workday.
        # Sundays/public holidays are already present explicitly in the diary sheets.
        for day in range(1, month_days + 1):
            if day not in day_map:
                day_map[day] = AttendanceRow(
                    employee_id=employee_id,
                    date=datetime(2026, 2, day).date(),
                    status="Absent",
                    start=None,
                    end=None,
                    work_hours=0.0,
                    ot_hours=0.0,
                    site_project=None,
                    transport=None,
                    remarks="NO WORK",
                )

        records.extend(day_map[day] for day in sorted(day_map))

    return records


def build_lookup_values(workbook_path: Path) -> dict[str, list[str]]:
    wb = load_workbook(workbook_path, data_only=True)
    sites: set[str] = set()
    transports: set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name not in PRIMARY_SHEETS:
            continue

        ws = wb[sheet_name]
        for row_idx in range(10, 38):
            raw_site = ws.cell(row_idx, 8).value
            raw_transport = ws.cell(row_idx, 12).value

            site = str(raw_site).strip() if raw_site not in (None, "", " ") else ""
            if site:
                sites.add(site)

            transport = ""
            if raw_transport not in (None, "", " ", "-", "—"):
                # The workbook uses compact transport codes in the diary column.
                transport = "MRT"
            if transport:
                transports.add(transport)

    return {
        "SiteProject": sorted(sites),
        "Transport": sorted(transports),
    }


def normalize_connection_string(connection_string: str) -> str:
    normalized = connection_string
    normalized = re.sub(r"User Id\s*=", "UID=", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"UserID\s*=", "UID=", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"Password\s*=", "PWD=", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"Encrypt\s*=\s*True", "Encrypt=yes", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"Encrypt\s*=\s*False", "Encrypt=no", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"TrustServerCertificate\s*=\s*True", "TrustServerCertificate=yes", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"TrustServerCertificate\s*=\s*False", "TrustServerCertificate=no", normalized, flags=re.IGNORECASE)
    if "DRIVER=" in normalized.upper():
        return normalized
    return f"DRIVER={{ODBC Driver 17 for SQL Server}};{normalized}"


def main() -> int:
    parser = argparse.ArgumentParser(description="Import February 2026 attendance from the diary workbook.")
    parser.add_argument("--workbook", required=True, type=Path, help="Path to FEB2026_ALL_WORKERS_DiaryReports_1.xlsx")
    parser.add_argument("--connection-string", default=os.environ.get("HRPAYROLL_CONNECTION_STRING"), help="SQL Server connection string")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print summary without writing to SQL Server")
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")
    if not args.connection_string:
        raise SystemExit("Missing connection string. Set HRPAYROLL_CONNECTION_STRING or pass --connection-string.")

    conn = pyodbc.connect(normalize_connection_string(args.connection_string), autocommit=False, timeout=30)
    cur = conn.cursor()

    employee_map: dict[str, int] = {}
    for employee_id, first_name, last_name in cur.execute(
        "SELECT Id, FirstName, LastName FROM SLE_Employees WHERE Status = 'Active'"
    ):
        employee_map[normalize_text(f"{first_name} {last_name}")] = int(employee_id)

    records = build_records(args.workbook, employee_map)
    lookup_values = build_lookup_values(args.workbook)
    if not records:
        raise SystemExit("No attendance rows were parsed from the workbook.")

    print(f"Parsed {len(records)} attendance rows from {args.workbook.name}")
    for rec in records[:5]:
        print(f"  emp={rec.employee_id} date={rec.date} status={rec.status} start={rec.start} end={rec.end} work={rec.work_hours} ot={rec.ot_hours} site={rec.site_project} transport={rec.transport} remarks={rec.remarks}")
    if len(records) > 5:
        print(f"  ... {len(records) - 5} more rows")

    if args.dry_run:
        print("Lookup values to seed:")
        for category, values in lookup_values.items():
            print(f"  {category}: {values}")
        return 0

    employee_ids = sorted({rec.employee_id for rec in records})
    month_start = datetime(2026, 2, 1).date()
    month_end = datetime(2026, 2, 28).date()

    print("Deleting existing February 2026 attendance for imported employees...")
    cur.execute(
        f"""
        DELETE FROM SLE_Attendances
        WHERE EmployeeId IN ({",".join("?" for _ in employee_ids)})
          AND [Date] >= ? AND [Date] <= ?
        """,
        employee_ids + [month_start, month_end],
    )

    now = datetime.utcnow()
    insert_sql = """
        INSERT INTO SLE_Attendances
            (EmployeeId, [Date], CheckIn, CheckOut, WorkHours, OTHours, SiteProject, Transport, Status, Remarks, CreatedAt, UpdatedAt)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    inserted = 0
    for rec in records:
        cur.execute(
            insert_sql,
            rec.employee_id,
            rec.date,
            rec.start,
            rec.end,
            rec.work_hours,
            rec.ot_hours,
            rec.site_project,
            rec.transport,
            rec.status,
            rec.remarks,
            now,
            now,
        )
        inserted += 1

    conn.commit()
    print(f"Inserted {inserted} attendance rows.")

    lookup_sql = """
        SELECT TOP 1 Id, Name, IsActive, SortOrder
        FROM SLE_AttendanceLookups
        WHERE Category = ? AND Name = ?
    """
    insert_lookup_sql = """
        INSERT INTO SLE_AttendanceLookups (Category, Name, IsActive, SortOrder, CreatedAt, UpdatedAt)
        VALUES (?, ?, 1, ?, ?, ?)
    """
    now = datetime.utcnow()
    print("Seeding lookup values...")
    for category, values in lookup_values.items():
        for idx, name in enumerate(values, start=1):
            existing = cur.execute(lookup_sql, category, name).fetchone()
            if existing is None:
                cur.execute(insert_lookup_sql, category, name, idx, now, now)
    conn.commit()

    # Quick verification output.
    for employee_id in employee_ids:
        count = cur.execute(
            "SELECT COUNT(*) FROM SLE_Attendances WHERE EmployeeId = ? AND [Date] >= ? AND [Date] <= ?",
            employee_id,
            month_start,
            month_end,
        ).fetchone()[0]
        print(f"  employee {employee_id}: {count} rows")

    for category, values in lookup_values.items():
        print(f"  {category}: {', '.join(values) if values else 'none'}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
